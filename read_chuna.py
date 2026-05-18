"""
구글 시트(추나현황) → 추나현황.csv 자동 업데이트 스크립트

실행 시 구글 시트에서 최신 xlsx를 자동 다운로드 후 파싱.
(구글 시트를 "링크가 있는 모든 사용자 - 뷰어" 공유 필요)

셀 구조 (이문환/방민준 시트):
  열A: datetime객체(날짜) 또는 문자열(기간/결산)
  열B: "당일예약 수" / "건보침" / "건보추나" / "상담" 등
  열C: 값
  열D: "ta침" / "ta추나" / "입원침" 등
  열E: 값
"""

import sys, csv, subprocess
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests"])
    import requests

SHEET_ID  = "1QTOX85QDWBKERfaVIXs071C63XMGaW9n9-rUcDBBDRY"
XLSX_URL  = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
XLSX_PATH = Path(r"C:\Users\하슬라한의원\한의원지표\data\추나시트.xlsx")
CSV_PATH  = Path(r"C:\Users\하슬라한의원\한의원지표\data\추나현황.csv")
DOCTORS   = ["노왕식", "이문환", "방민준"]


def download_xlsx():
    """구글 시트에서 최신 xlsx 다운로드 (공개 공유 필요)"""
    print(f"구글 시트 다운로드 중...")
    resp = requests.get(XLSX_URL, timeout=30)
    if resp.status_code == 200:
        XLSX_PATH.write_bytes(resp.content)
        print(f"  저장 완료: {XLSX_PATH}")
        return True
    else:
        print(f"  [경고] 다운로드 실패 (HTTP {resp.status_code}) → 기존 파일 사용")
        return False


def to_int(v):
    try:
        return int(float(v)) if v is not None and str(v).strip() != "" else None
    except:
        return None


def parse_sheet(ws):
    """시트에서 일별 건보추나/TA추나 추출 (결산 행 제외)"""
    records = {}        # 날짜 → {건보추나, TA추나}
    current_date = None
    in_summary   = False  # 결산 구간 플래그

    for row in ws.iter_rows():
        if len(row) < 2:
            continue

        col_a = row[0].value
        col_b = str(row[1].value).strip() if row[1].value is not None else ""

        # ── 날짜 판별 ─────────────────────────────────────────────────────────
        if isinstance(col_a, datetime):
            # 실제 날짜 셀 → 일별 구간 시작
            current_date = col_a.strftime("%Y-%m-%d")
            in_summary   = False
            continue

        if col_a is not None and str(col_a).strip() != "":
            a_str = str(col_a).strip()
            # 요일 한 글자('월','화','수','목','금','토','일')는 무시
            if a_str in {"월", "화", "수", "목", "금", "토", "일"}:
                continue
            # 기간 범위·결산 문자열 → 결산 구간
            in_summary = True
            continue

        # ── 건보추나 행 처리 ──────────────────────────────────────────────────
        if col_b == "건보추나" and not in_summary and current_date:
            geonbo = to_int(row[2].value) if len(row) > 2 else None
            ta     = to_int(row[4].value) if len(row) > 4 else None

            if geonbo is None and ta is None:
                continue

            # 같은 날짜가 이미 있으면 덮어쓰지 않음 (첫 번째 일별 값 우선)
            if current_date not in records:
                records[current_date] = {
                    "건보추나": geonbo if geonbo is not None else 0,
                    "TA추나":   ta     if ta     is not None else 0,
                }

    return [{"날짜": k, **v} for k, v in sorted(records.items())]


def load_csv():
    rows = []
    if CSV_PATH.exists():
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                rows.append(r)
    return rows


def save_csv(rows):
    with open(CSV_PATH, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["날짜","원장명","건보추나","TA추나"], quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows(rows)


def main():
    download_xlsx()
    if not XLSX_PATH.exists():
        print(f"[오류] 파일 없음: {XLSX_PATH}")
        sys.exit(1)

    print(f"xlsx 로드: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"시트: {wb.sheetnames}\n")

    sheet_data = {}
    for doc in DOCTORS:
        if doc not in wb.sheetnames:
            print(f"[경고] '{doc}' 시트 없음")
            continue

        records = parse_sheet(wb[doc])

        # 이문환: 4/27 이전 제거
        if doc == "이문환":
            records = [r for r in records if r["날짜"] >= "2026-04-27"]

        # 0,0 행 제거
        records = [r for r in records if not (r["건보추나"] == 0 and r["TA추나"] == 0)]

        sheet_data[doc] = records
        print(f"[{doc}] {len(records)}건")
        for r in records:
            print(f"  {r['날짜']}  건보추나={r['건보추나']:2d}  TA추나={r['TA추나']:2d}")
        print()

    # CSV 업데이트: 대상 원장만 교체, 노왕식은 그대로
    existing = load_csv()
    kept = [r for r in existing if r["원장명"] not in sheet_data]
    new_rows = []
    for doc, records in sheet_data.items():
        for rec in records:
            new_rows.append({"날짜": rec["날짜"], "원장명": doc,
                             "건보추나": rec["건보추나"], "TA추나": rec["TA추나"]})

    all_rows = sorted(kept + new_rows, key=lambda x: (x["날짜"], x["원장명"]))
    save_csv(all_rows)
    print(f"CSV 저장 완료: {len(all_rows)}행 → {CSV_PATH}")


if __name__ == "__main__":
    main()
